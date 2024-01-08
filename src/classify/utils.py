import pandas as pd
import datetime
import openpyxl
def get_time_interval(time):
    y = pd.to_datetime(time)[0].year
    m = pd.to_datetime(time)[0].month
    d = pd.to_datetime(time)[0].day

    start1 = datetime.datetime(y,m,d,0,0,0)
    end1   = datetime.datetime(y,m,d,12,0,0)

    start2 = datetime.datetime(y,m,d,12,0,0)
    end2   = datetime.datetime(y,m,d+1,0,0,0)

    return start1, end1, start2, end2
def get_date_str(time):
    current_date = pd.Timestamp(f'{pd.to_datetime(time.data[0]).year}-{pd.to_datetime(time.data[0]).month}-{pd.to_datetime(time.data[0]).day}')
    ymd = 10000*current_date.year + 100*current_date.month + current_date.day
    return ymd
def append_to_excel( path_to_excel , date, warm_cloud_column,trade_cu_column, cold_cloud_column, no_cloud_column, time_steps_column ):
        # Open the workbook
        workbook = openpyxl.load_workbook(path_to_excel)

        # Select the sheet to work with
        sheet = workbook['Tabelle1']

        # Get the next available row
        next_row = sheet.max_row + 1

        # Write the data to the sheet
        sheet.cell(row=next_row, column=2).value = warm_cloud_column
        sheet.cell(row=next_row, column=1).value = date
        sheet.cell(row=next_row, column=3).value = trade_cu_column
        sheet.cell(row=next_row, column=4).value = cold_cloud_column
        sheet.cell(row=next_row, column=5).value = no_cloud_column
        sheet.cell(row=next_row, column=6).value = time_steps_column



        # Save the changes to the workbook
        workbook.save(path_to_excel)
